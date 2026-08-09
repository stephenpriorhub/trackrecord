#!/usr/bin/env python3
"""
Build data/warRoomOwners.json — the verified Bryan/Karim owner map for The War Room.

Source: the manually-maintained War Room track-record workbook that editorial exports
weekly ("War Room (Bryan and Karim Only) - Weekly Report <date>.xlsx"). Every trade row in
it carries a single-letter B/K column: Bryan and Karim never co-own a War Room pick, so
this workbook is the authority on who made each one.

Every sheet with a B/K column is harvested (they are overlapping cuts of the same trade
log — YTD, L12, Karim Only, and so on — and some hold rows the others don't). Columns are
located by header name, not position, because the cuts order them differently.

Output shape, most specific lookup first:
  occ    "TICKER|YYMMDD|C/P|STRIKE" -> owner   exact option contract
  dated  "TICKER@YYYY-MM-DD"        -> owner   underlying + open date; bridges the sheet
                                               listing a position by company name/ticker
                                               while Airtable lists it by contract
  ticker "TICKER"                   -> owner   last owner to trade it; ambiguous, last resort

Usage:  python3 scripts/build-war-room-owners.py <workbook.xlsx> > data/warRoomOwners.json
Requires openpyxl.
"""
import sys
import re
import json
import datetime
from collections import defaultdict, Counter

import openpyxl

OCC = re.compile(r'^([A-Z]+)(\d{6})([CP])(\d{6,8})')
OWNERS = {'B': 'bryan', 'K': 'karim'}


def norm(sym):
    """Normalize a symbol to an OCC key, or a bare ticker."""
    s = re.sub(r'[^A-Z0-9]', '', str(sym).upper())
    if not s:
        return None
    m = OCC.match(s)
    if m:
        return f"{m.group(1)}|{m.group(2)}|{m.group(3)}|{int(m.group(4))}"
    return s


def find_header(ws):
    """Locate the header row and the columns we need. Returns None if the sheet has no B/K."""
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        cells = {str(c).strip().lower(): i for i, c in enumerate(row) if c is not None}
        if 'b/k' in cells and ('symbol' in cells or 'name' in cells):
            return {
                'bk': cells['b/k'],
                'symbol': cells.get('symbol'),
                'name': cells.get('name'),
                'open': cells.get('open date'),
                'row': row,
            }
    return None


def harvest(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    occ, dated, tick = defaultdict(list), defaultdict(list), defaultdict(list)
    seq = 0
    sheets_used = []

    for ws in wb.worksheets:
        hdr = find_header(ws)
        if not hdr:
            continue
        sheets_used.append(ws.title)
        started = False
        for row in ws.iter_rows(values_only=True):
            if not started:
                started = row is hdr['row'] or list(row) == list(hdr['row'])
                continue
            if not any(c is not None for c in row):
                continue

            def cell(i):
                return row[i] if i is not None and i < len(row) else None

            bk = cell(hdr['bk'])
            owner = OWNERS.get(str(bk).strip().upper()) if bk is not None else None
            if not owner:
                continue
            # A handful of rows leave Symbol blank and carry the contract in Name instead.
            raw = cell(hdr['symbol']) or cell(hdr['name'])
            if not raw:
                continue

            od = cell(hdr['open'])
            day = od.strftime('%Y-%m-%d') if isinstance(od, datetime.datetime) else None
            seq += 1
            for part in str(raw).split(','):
                k = norm(part)
                if not k:
                    continue
                if '|' in k:
                    occ[k].append((seq, owner))
                    under = k.split('|')[0]
                else:
                    # Guard against free-text names ("Viking Therapeut") normalizing to junk.
                    if len(k) > 6 or not k.isalpha():
                        continue
                    under = k
                    tick[k].append((seq, owner))
                if day:
                    dated[f"{under}@{day}"].append((seq, owner))

    return occ, dated, tick, seq, sheets_used


def collapse(m, label):
    out, conflicts = {}, 0
    for k, entries in m.items():
        if len({o for _, o in entries}) > 1:
            conflicts += 1
        # Sheets are chronological; the latest row wins.
        out[k] = max(entries, key=lambda e: e[0])[1]
    print(f"{label}: {len(out)} keys, {conflicts} conflicting", file=sys.stderr)
    return out


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    path = sys.argv[1]
    occ, dated, tick, rows, sheets = harvest(path)
    data = {
        'source': path.split('/')[-1],
        'generatedFrom': f'{rows} rows across {len(sheets)} sheets: {", ".join(sheets)}',
        'occ': collapse(occ, 'occ'),
        'dated': collapse(dated, 'ticker@openDate'),
        'ticker': collapse(tick, 'ticker'),
    }
    print(Counter(list(data['occ'].values()) + list(data['ticker'].values())), file=sys.stderr)
    json.dump(data, sys.stdout, separators=(',', ':'), sort_keys=True)


if __name__ == '__main__':
    main()
