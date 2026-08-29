"""
Read a publication's published track-record spreadsheet into normalized JSON.

Why Python: the sheets are .xlsx and only the raw cell values are trustworthy.
Google's CSV export hands back FORMATTED text — "100%" where the real figure is
1.0102, "$11.00 " where it is 11 — so a CSV import would publish rounded returns.
openpyxl reads the underlying numbers. (scripts/build-war-room-owners.py reads
the War Room workbook the same way, for the same reason.)

The output is deliberately dumb: one flat record per closed trade, with the
sheet's own return carried alongside so the importer can check its arithmetic
against the publisher's rather than assume they agree.

  python3 scripts/extract-sheet-trades.py --out /tmp/trades.json
  python3 scripts/extract-sheet-trades.py --local dpl.xlsx --out /tmp/dpl.json

Needs openpyxl:  python3 -m pip install openpyxl
"""

import argparse
import json
import re
import sys
import urllib.request
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python3 -m pip install openpyxl")

# The three published track records, by Google Sheet id.
SHEETS = {
    "DPL": "1C7_QAbPU1QUKT_h4dm3XvAti9-sJi-CquoCFfCyaPEg",
    "PSU": "1k9XZCI5MxkA5gswfo9eUipT6UBdQUDAYy7mo80kdSGM",
    "NBS": "1cbdDD508riq_TC9whm3G5csN4dZdblhfbbENHULSqwM",
}

# The Portfolio column is the routing key, NOT the file a row arrived in: the
# DPL workbook carries two Profit Surge Trader trades and the PSU workbook
# carries six Daily Profits Live ones. Importing by filename would file those
# under the wrong publication.
PORTFOLIO_CODES = {
    # long form, used by the main row schema
    "sector strike": "NBS",
    "daily profits live": "DPL",
    "profit surge trader": "PSU",
    # short form, used by the fill row schema
    "ss": "NBS",
    "dpl": "DPL",
    "psu": "PSU",
}

TAB = "Since Inception"

# ---- main row schema -------------------------------------------------------
# One row per completed trade.
#
# The basis is COST and RETURN %, not the price columns. Three things forced
# that, in order of how badly each would have misreported the record:
#
#   - Entry/Exit Price are the first and last print. A trade scaled in and out
#     shows entry 6.50 / exit 3.10 on a +85.54% winner.
#   - Avg Buy / Avg Sell fix that for a single-leg trade, but on a vertical or
#     butterfly they average across BOTH legs and are not the net debit. 272
#     rows disagreed with the published return on this basis alone.
#   - Return $ / Cost reproduces the sheet's own Return % on 3,791 of 3,801
#     testable rows, spreads included.
#
# So Return % is taken verbatim and the entry is derived from Cost. That makes
# our averages identical to the ones the publisher prints, which is the only
# reconciliation that actually matters.
M = dict(status=0, symbol=1, size=2, opened=3, closed=4, entry=7, exit=8,
         ret_dollars=9, ret=10, avg_buy=11, avg_sell=12, expiry=14, strike=15,
         type=16, side=17, spread=18, cost=19, portfolio=23)

# ---- fill row schema -------------------------------------------------------
# A second table appended below the first, in the layout of the "MTD" / "This
# Month" tabs: one row per closing FILL, carrying the full OCC symbol. Its
# header lives on those tabs, not above these rows.
F = dict(status=0, occ=1, name=2, fill_price=3, qty=4, closed=5, close_time=6,
         running=7, avg_buy=8, avg_sell=9, ret=10, portfolio=11, opened=12)

OCC_RE = re.compile(r"^([A-Z]+)(\d{6})([CP])(\d+)$")


def as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def as_num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def txt(v):
    return "" if v is None else str(v).strip()


def parse_occ(sym):
    """OCC symbol -> contract. Strikes here are NOT zero-padded to 8 digits."""
    m = OCC_RE.match(sym.strip().upper())
    if not m:
        return None
    root, ymd, right, strike = m.groups()
    try:
        expiry = datetime.strptime(ymd, "%y%m%d").date().isoformat()
    except ValueError:
        return None
    return {
        "underlying": root,
        "expiry": expiry,
        "right": "CALL" if right == "C" else "PUT",
        "strike": int(strike) / 1000.0,
    }


def infer_kind(r):
    """
    Option or stock, for a row that does not say.

    Read off the sheet's OWN arithmetic rather than guessed: Return $ is the
    price move times the contract multiplier, so whichever of 100 or 1
    reproduces it is the multiplier this row was written with. Ties and
    unreadable rows fall to OPTION, which is what these services trade.
    """
    entry, exit_ = as_num(r[M["entry"]]), as_num(r[M["exit"]])
    ret_d = as_num(r[M["ret_dollars"]])
    if entry is None or exit_ is None or ret_d is None:
        return "OPTION"
    move = exit_ - entry
    if abs(move) < 1e-9:
        return "OPTION"
    for mult, kind in ((100, "OPTION"), (1, "STOCK")):
        for size in (1, as_num(r[M["size"]]) or 1):
            if abs(move * mult * size - ret_d) < 0.51:
                return kind
    return "OPTION"


def main_row(r, n):
    """A completed trade in the primary schema, or None."""
    status = txt(r[M["status"]]).upper()
    if status not in ("WIN", "LOSS", "BE"):
        return None
    symbol = txt(r[M["symbol"]]).upper()
    opened, closed = as_date(r[M["opened"]]), as_date(r[M["closed"]])
    if not symbol or not opened or not closed:
        return None

    kind_raw = txt(r[M["type"]]).upper()
    if kind_raw in ("SHARE", "SHARES", "STOCK"):
        kind, right = "STOCK", None
    elif kind_raw == "CALL":
        kind, right = "OPTION", "CALL"
    elif kind_raw == "PUT":
        kind, right = "OPTION", "PUT"
    elif kind_raw == "OPTION":
        # The sheet says it is an option but not which side. Recorded as
        # unknown; the importer must not guess.
        kind, right = "OPTION", None
    elif not kind_raw:
        # A third variant: rows flagged "NO SYNC", entered by hand when the
        # broker feed missed a trade. They carry Status, dates, Entry/Exit Price
        # and Return %, but no Type, Cost or averages. Real trades, so they are
        # imported — the contract simply goes unstated, like most of the rest.
        kind, right = infer_kind(r), None
    else:
        # "NO SELL ALERT", "custom return- roll" and friends are not trades.
        return "not a trade: type " + repr(kind_raw)

    size = as_num(r[M["size"]])
    units = int(size) if size and size >= 1 else 1

    # A spread's legs are not recoverable from one row — the Strike cell for a
    # butterfly reads "$3,780.00 / $3,800.00 / $3,820.00". It is carried as one
    # synthetic position priced at the net, and labelled as the spread it was.
    spread = txt(r[M["spread"]]).upper() or "SINGLE"
    multi_leg = spread not in ("SINGLE", "")

    cost = as_num(r[M["cost"]])
    ret_pct = as_num(r[M["ret"]])
    ret_dollars = as_num(r[M["ret_dollars"]])

    # Return % as published, wherever it exists. Recomputing it from Return $
    # is only a fallback, because the sheet's own figure is the one that feeds
    # the averages readers have already seen.
    if ret_pct is None and ret_dollars is not None and cost:
        ret_pct = ret_dollars / cost
    if ret_pct is None:
        return "no return recorded"

    multiplier = 100 if kind == "OPTION" else 1
    entry = cost / (units * multiplier) if cost and units else None
    if entry is None or entry <= 0:
        # No cost basis: fall back to the averages, which are exact for a
        # single-leg trade and are all this row offers.
        entry = as_num(r[M["avg_buy"]]) or as_num(r[M["entry"]])
    if entry is None or entry <= 0:
        return "no entry price"
    exit_ = entry * (1 + ret_pct)

    # Only a plain numeric strike is usable; a spread's compound string is not.
    strike = as_num(r[M["strike"]])
    return {
        "row": n,
        "rowKind": "main",
        "status": status,
        "portfolio": txt(r[M["portfolio"]]),
        "symbol": symbol,
        "openedAt": opened,
        "closedAt": closed,
        "entry": entry,
        "exit": exit_,
        "units": units,
        "kind": kind,
        "right": None if multi_leg else right,
        "strike": None if multi_leg else strike,
        "expiry": None if multi_leg else as_date(r[M["expiry"]]),
        "side": "SELL" if txt(r[M["side"]]).upper() == "SHORT" else "BUY",
        "spread": spread,
        "multiLeg": multi_leg,
        "sheetReturn": ret_pct,
        "occ": None,
    }


def fill_row(r, n):
    """A closing fill in the secondary schema, or None."""
    occ_raw = txt(r[F["occ"]]).upper()
    contract = parse_occ(occ_raw)
    if not contract:
        return None

    # Running position 0 means this fill CLOSED the trade. A scale-out writes
    # one row per exit with the running position falling to zero, and the final
    # row's weighted averages already cover the whole trade — so taking only
    # the zero rows records each trade once, complete, with no double count.
    running = as_num(r[F["running"]])
    if running is None or abs(running) > 1e-9:
        return "partial exit (running position not zero)"

    opened, closed = as_date(r[F["opened"]]), as_date(r[F["closed"]])
    entry, exit_ = as_num(r[F["avg_buy"]]), as_num(r[F["avg_sell"]])
    if not opened or not closed:
        return "no open or close date"
    if entry is None or exit_ is None or entry <= 0:
        return "no entry or exit price"

    # Same rule as the main schema: the published Return % wins. On 13 rows the
    # weighted averages on the closing fill do not span the whole trade (an
    # earlier scale-out is missing from them), and recomputing would print a
    # return the publisher never showed — INOD as -100% where the record says
    # -81.45%.
    ret = as_num(r[F["ret"]])
    if ret is not None:
        exit_ = entry * (1 + ret)

    qty = as_num(r[F["qty"]])
    return {
        "row": n,
        "rowKind": "fill",
        "status": txt(r[F["status"]]) or "Sell",
        "portfolio": txt(r[F["portfolio"]]),
        "symbol": contract["underlying"],
        "openedAt": opened,
        "closedAt": closed,
        "entry": entry,
        "exit": exit_,
        "units": int(qty) if qty and qty >= 1 else 1,
        "kind": "OPTION",
        "right": contract["right"],
        "strike": contract["strike"],
        "expiry": contract["expiry"],
        "side": "BUY",
        "spread": "SINGLE",
        "multiLeg": False,
        # This schema states the contract in full, so the return is checkable
        # and is recomputed rather than trusted blind.
        "sheetReturn": as_num(r[F["ret"]]),
        "occ": occ_raw,
    }


def read_workbook(path_or_bytes, source):
    wb = openpyxl.load_workbook(path_or_bytes, data_only=True, read_only=True)
    if TAB not in wb.sheetnames:
        sys.exit(f"{source}: no {TAB!r} tab (found {wb.sheetnames})")
    rows, skipped = [], []
    for n, r in enumerate(wb[TAB].iter_rows(values_only=True), start=1):
        if n == 1 or not any(v is not None for v in r):
            continue
        # Pad short rows so a fixed column index never raises.
        r = tuple(r) + (None,) * 48
        # A row is one schema or the other: None means "not this shape, try the
        # next", while a string means "this shape, but unusable" — and every one
        # of those is reported rather than silently dropped.
        rec = main_row(r, n)
        if rec is None:
            rec = fill_row(r, n)
        if rec is None:
            if txt(r[1]):
                skipped.append((n, txt(r[1]), "unrecognised row"))
            continue
        if isinstance(rec, str):
            skipped.append((n, txt(r[1]), rec))
            continue
        rec["source"] = source
        rows.append(rec)
    wb.close()
    return rows, skipped


def fetch(sheet_id):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("--local", nargs="*", default=None,
                    help="local .xlsx files instead of downloading")
    args = ap.parse_args()

    import io as _io
    all_rows, report, all_skips = [], [], []
    sources = (
        [(p, p) for p in args.local]
        if args.local
        else [(code, sid) for code, sid in SHEETS.items()]
    )
    for label, ref in sources:
        data = ref if args.local else _io.BytesIO(fetch(ref))
        rows, skipped = read_workbook(data, label)
        all_rows.extend(rows)
        report.append((label, len(rows), len(skipped)))
        all_skips.extend((label, *sk) for sk in skipped)

    # Route every row by its own Portfolio cell.
    unmapped = {}
    for rec in all_rows:
        code = PORTFOLIO_CODES.get(rec["portfolio"].strip().lower())
        rec["pubCode"] = code
        if code is None:
            unmapped[rec["portfolio"]] = unmapped.get(rec["portfolio"], 0) + 1

    with open(args.out, "w") as fh:
        json.dump(all_rows, fh, indent=1)

    print(f"{'sheet':6} {'trades':>7} {'skipped':>8}")
    for label, n, sk in report:
        print(f"{label:6} {n:7} {sk:8}")
    counts = {}
    for rec in all_rows:
        k = rec["pubCode"] or f"?? {rec['portfolio']!r}"
        counts[k] = counts.get(k, 0) + 1
    print("\nby publication:")
    for k in sorted(counts, key=lambda x: -counts[x]):
        print(f"  {k:28} {counts[k]:6}")
    if unmapped:
        print("\nUNMAPPED portfolio labels (these rows will NOT import):")
        for k, v in sorted(unmapped.items(), key=lambda x: -x[1]):
            print(f"  {k!r:40} {v}")

    if all_skips:
        reasons = {}
        for _, _, _, why in all_skips:
            reasons[why] = reasons.get(why, 0) + 1
        print(f"\nskipped {len(all_skips)} rows:")
        for why, n in sorted(reasons.items(), key=lambda x: -x[1]):
            print(f"  {n:5}  {why}")
        print("  first few:")
        for label, n, sym, why in all_skips[:8]:
            print(f"    {label:>10} row {n:<6} {sym:<10} {why}")
    print(f"\nwrote {len(all_rows)} trades -> {args.out}")


if __name__ == "__main__":
    main()
